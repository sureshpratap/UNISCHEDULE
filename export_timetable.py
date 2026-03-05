"""
University Timetable Generator — Excel Export Script
======================================================
Run: python export_timetable.py
Generates a sample timetable XLSX with proper formatting.
Plug in your own data in the DATA section below.
"""

import json
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Colours ───────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1a1d27")
SUBHEADER_FILL= PatternFill("solid", fgColor="2e3352")
DAY_FILL      = PatternFill("solid", fgColor="21253a")
LUNCH_FILL    = PatternFill("solid", fgColor="3d3600")
FREE_FILL     = PatternFill("solid", fgColor="161922")
WHITE         = Font(color="FFFFFF", bold=True, name="Arial", size=10)
MUTED         = Font(color="8892b0", name="Arial", size=9)
TITLE_FONT    = Font(color="FFFFFF", bold=True, name="Arial", size=13)
GOLD          = Font(color="f5c542", bold=True, name="Arial", size=9)

SUBJECT_COLORS = [
    "6c63ff","22d3ee","f5c542","34d399","f87171","fb923c",
    "a78bfa","38bdf8","4ade80","facc15","f472b6","818cf8",
]

thin  = Side(style="thin",  color="2e3352")
thick = Side(style="medium", color="6c63ff")
THIN_BORDER  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
THICK_BORDER = Border(left=thick, right=thick, top=thick, bottom=thick)

# ── SAMPLE DATA  (replace with your actual data) ──────────────────────────────
DATA = {
    "university": "Shiv Nadar University",
    "academic_year": "2025-26  |  Odd Semester",
    "departments": [
        {"id": "d1", "name": "Computer Science & Engineering"},
    ],
    "subjects": [
        {"id": "s1", "name": "Data Structures",      "hours_per_class": 1, "hours_per_week": 4, "is_lab": False},
        {"id": "s2", "name": "Operating Systems",     "hours_per_class": 1, "hours_per_week": 4, "is_lab": False},
        {"id": "s3", "name": "DS Lab",                "hours_per_class": 3, "hours_per_week": 3, "is_lab": True},
        {"id": "s4", "name": "OS Lab",                "hours_per_class": 3, "hours_per_week": 3, "is_lab": True},
        {"id": "s5", "name": "Tutorial",              "hours_per_class": 2, "hours_per_week": 2, "is_lab": False},
    ],
    "faculty": [
        {"id": "f1", "name": "Dr. Priya Sharma",  "dept_id": "d1", "subjects": ["Data Structures", "Tutorial"]},
        {"id": "f2", "name": "Prof. Arjun Mehta", "dept_id": "d1", "subjects": ["Operating Systems"]},
        {"id": "f3", "name": "Ms. Kavya Nair",    "dept_id": "d1", "subjects": ["DS Lab", "OS Lab"]},
    ],
    "timing": {
        "class_days": ["Monday","Tuesday","Wednesday","Thursday","Friday"],
        "lunch_start": "13:00",
        "lunch_end":   "14:00",
        "slots": [
            {"id":"sl1","label":"Period 1","start":"09:00","end":"10:00"},
            {"id":"sl2","label":"Period 2","start":"10:00","end":"11:00"},
            {"id":"sl3","label":"Period 3","start":"11:00","end":"12:00"},
            {"id":"sl4","label":"Period 4","start":"12:00","end":"13:00"},
            {"id":"sl5","label":"Period 5","start":"14:00","end":"15:00"},
            {"id":"sl6","label":"Period 6","start":"15:00","end":"16:00"},
        ],
    },
    "sections": [
        {
            "id": "sec1", "name": "CS-3A", "room": "LH-301", "dept_id": "d1",
            "timetable": {
                "Monday":    {"sl1":"s1/f1","sl2":"s2/f2","sl3":None,"sl4":None,"sl5":"s1/f1","sl6":None},
                "Tuesday":   {"sl1":None,"sl2":None,"sl3":"s3/f3","sl4":"s3/f3","sl5":"s3/f3","sl6":None},
                "Wednesday": {"sl1":"s2/f2","sl2":"s5/f1","sl3":"s5/f1","sl4":None,"sl5":"s1/f1","sl6":"s2/f2"},
                "Thursday":  {"sl1":None,"sl2":None,"sl3":"s4/f3","sl4":"s4/f3","sl5":"s4/f3","sl6":None},
                "Friday":    {"sl1":"s1/f1","sl2":"s2/f2","sl3":None,"sl4":None,"sl5":None,"sl6":None},
            }
        },
        {
            "id": "sec2", "name": "CS-3B", "room": "LH-302", "dept_id": "d1",
            "timetable": {
                "Monday":    {"sl1":"s2/f2","sl2":"s1/f1","sl3":None,"sl4":None,"sl5":"s2/f2","sl6":None},
                "Tuesday":   {"sl1":"s1/f1","sl2":None,"sl3":None,"sl4":None,"sl5":None,"sl6":None},
                "Wednesday": {"sl1":None,"sl2":None,"sl3":"s4/f3","sl4":"s4/f3","sl5":"s4/f3","sl6":None},
                "Thursday":  {"sl1":"s1/f1","sl2":"s5/f1","sl3":"s5/f1","sl4":None,"sl5":"s2/f2","sl6":None},
                "Friday":    {"sl1":None,"sl2":None,"sl3":"s3/f3","sl4":"s3/f3","sl5":"s3/f3","sl6":None},
            }
        },
    ]
}

# ── Helper lookups ─────────────────────────────────────────────────────────────
def build_lookups(data):
    subjects = {s["id"]: s for s in data["subjects"]}
    faculty  = {f["id"]: f for f in data["faculty"]}
    return subjects, faculty

def is_lunch(slot, timing):
    return slot["start"] < timing["lunch_end"] and slot["end"] > timing["lunch_start"]

def subject_color(subject_name, subjects):
    names = [s["name"] for s in subjects]
    idx   = names.index(subject_name) if subject_name in names else 0
    return SUBJECT_COLORS[idx % len(SUBJECT_COLORS)]

def apply_cell(ws, row, col, value, font=None, fill=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font      = font
    if fill:   c.fill      = fill
    if align:  c.alignment = align
    if border: c.border    = border
    return c

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── Build workbook ─────────────────────────────────────────────────────────────
def build_workbook(data):
    wb = Workbook()
    subjects, faculty = build_lookups(data)
    slots   = sorted(data["timing"]["slots"], key=lambda x: x["start"])
    c_days  = data["timing"]["class_days"]
    timing  = data["timing"]

    # ── Cover / Summary sheet ─────────────────────────────────────────────────
    cover = wb.active
    cover.title = "Summary"
    cover.sheet_view.showGridLines = False
    cover.column_dimensions["A"].width = 30
    cover.column_dimensions["B"].width = 40

    def cov(r, c, v, **kw): return apply_cell(cover, r, c, v, **kw)

    cover.row_dimensions[1].height = 40
    cover.merge_cells("A1:B1")
    cov(1,1, data["university"], font=Font(color="6c63ff",bold=True,name="Arial",size=16), fill=HEADER_FILL, align=CENTER)

    cover.row_dimensions[2].height = 22
    cover.merge_cells("A2:B2")
    cov(2,1, data["academic_year"], font=MUTED, fill=HEADER_FILL, align=CENTER)

    r = 4
    headers = [("Departments",len(data["departments"])),("Subjects",len(data["subjects"])),
               ("Faculty",len(data["faculty"])),("Sections",len(data["sections"])),
               ("Class Days",len(c_days)),("Periods/Day",len(slots))]
    for label, val in headers:
        cov(r, 1, label, font=Font(color="8892b0",name="Arial",size=10), fill=DAY_FILL, align=LEFT, border=THIN_BORDER)
        cov(r, 2, val,   font=Font(color="FFFFFF",bold=True,name="Arial",size=10), fill=DAY_FILL, align=CENTER, border=THIN_BORDER)
        cover.row_dimensions[r].height = 22
        r += 1

    r += 1
    cov(r,1,"Faculty & Subjects", font=Font(color="8b85ff",bold=True,name="Arial",size=11), fill=SUBHEADER_FILL, align=LEFT)
    cover.merge_cells(f"A{r}:B{r}")
    r += 1
    for f in data["faculty"]:
        cov(r,1, f["name"], font=Font(color="FFFFFF",name="Arial",size=10), fill=DAY_FILL, align=LEFT, border=THIN_BORDER)
        cov(r,2, ", ".join(f["subjects"]), font=MUTED, fill=DAY_FILL, align=LEFT, border=THIN_BORDER)
        cover.row_dimensions[r].height = 20
        r += 1

    # ── One sheet per section ─────────────────────────────────────────────────
    for sec in data["sections"]:
        dept = next((d for d in data["departments"] if d["id"] == sec["dept_id"]), {})
        ws = wb.create_sheet(title=sec["name"])
        ws.sheet_view.showGridLines = False

        # Column widths
        ws.column_dimensions["A"].width = 14  # day column
        for ci in range(2, len(slots)+2):
            ws.column_dimensions[get_column_letter(ci)].width = 18

        # ── Row 1: Section title ──────────────────────────────────────────────
        ws.merge_cells(f"A1:{get_column_letter(len(slots)+1)}1")
        ws.row_dimensions[1].height = 36
        title_str = f"{sec['name']}  |  Room: {sec['room']}  |  {dept.get('name','')}  |  {data['academic_year']}"
        apply_cell(ws,1,1, title_str,
                   font=Font(color="FFFFFF",bold=True,name="Arial",size=12),
                   fill=PatternFill("solid", fgColor="1a1d27"),
                   align=CENTER, border=THICK_BORDER)

        # ── Row 2: Slot headers ───────────────────────────────────────────────
        ws.row_dimensions[2].height = 30
        apply_cell(ws,2,1,"Day / Period",
                   font=WHITE, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)
        for ci, sl in enumerate(slots, 2):
            apply_cell(ws,2,ci, f"{sl['label']}\n{sl['start']}–{sl['end']}",
                       font=Font(color="8b85ff",bold=True,name="Arial",size=9),
                       fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)

        # ── Rows 3+: Day rows ────────────────────────────────────────────────
        for ri, day in enumerate(c_days, 3):
            ws.row_dimensions[ri].height = 48
            apply_cell(ws,ri,1, day,
                       font=Font(color="6c63ff",bold=True,name="Arial",size=10),
                       fill=DAY_FILL, align=CENTER, border=THIN_BORDER)

            for ci, sl in enumerate(slots, 2):
                if is_lunch(sl, timing):
                    apply_cell(ws,ri,ci,"🍽  LUNCH BREAK",
                               font=GOLD, fill=LUNCH_FILL, align=CENTER, border=THIN_BORDER)
                    continue

                cell_val = sec["timetable"].get(day, {}).get(sl["id"])
                if not cell_val:
                    apply_cell(ws,ri,ci,"—",
                               font=Font(color="2e3352",name="Arial",size=10),
                               fill=FREE_FILL, align=CENTER, border=THIN_BORDER)
                else:
                    sub_id, fac_id = cell_val.split("/")
                    sub = subjects.get(sub_id, {})
                    fac = faculty.get(fac_id, {})
                    sub_name = sub.get("name","")
                    fac_name = fac.get("name","")
                    is_lab   = sub.get("is_lab", False)
                    tag      = " [LAB]" if is_lab else ""
                    col_hex  = subject_color(sub_name, data["subjects"])

                    cell_text = f"{sub_name}{tag}\n{fac_name}"
                    bg = PatternFill("solid", fgColor=col_hex + "44"[:2] if len(col_hex)==6 else col_hex)
                    # Use a readable dark overlay
                    bg = PatternFill("solid", fgColor="1e2236")
                    apply_cell(ws,ri,ci, cell_text,
                               font=Font(color=col_hex, bold=True, name="Arial", size=9),
                               fill=bg, align=CENTER, border=THIN_BORDER)

        # ── Faculty load block below timetable ───────────────────────────────
        load_row = len(c_days) + 4
        ws.merge_cells(f"A{load_row}:{get_column_letter(len(slots)+1)}{load_row}")
        apply_cell(ws,load_row,1,"Faculty Teaching Load — This Section",
                   font=Font(color="8b85ff",bold=True,name="Arial",size=10),
                   fill=SUBHEADER_FILL, align=LEFT)
        ws.row_dimensions[load_row].height = 22
        load_row += 1

        fac_count = {}
        for day in c_days:
            for sl in slots:
                cv = sec["timetable"].get(day,{}).get(sl["id"])
                if cv:
                    _, fid = cv.split("/")
                    fac_count[fid] = fac_count.get(fid,0) + 1

        for fid, count in fac_count.items():
            f = faculty.get(fid, {})
            apply_cell(ws,load_row,1, f.get("name",""),
                       font=Font(color="FFFFFF",name="Arial",size=9), fill=DAY_FILL,
                       align=LEFT, border=THIN_BORDER)
            apply_cell(ws,load_row,2, f"{count} hrs/wk in this section",
                       font=MUTED, fill=DAY_FILL, align=LEFT, border=THIN_BORDER)
            ws.row_dimensions[load_row].height = 18
            load_row += 1

    return wb


if __name__ == "__main__":
    wb = build_workbook(DATA)
    out = "university_timetable.xlsx"
    wb.save(out)
    print(f"✅ Timetable saved to {out}")
    print("   Sheets:", wb.sheetnames)
